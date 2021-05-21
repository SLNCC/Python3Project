# !/usr/bin/python3
# -*- coding: UTF-8 -*-

__author__ = 'joedd'

import  openpyxl

from openpyxl.utils.cell import get_column_letter,column_index_from_string
# openpyxl 模块让 Python 程序能读取和修改 Excel 电子表格文件。
"""
    openpyxl.load_workbook()函数接受文件名，返回一个 workbook 数据类型的值。
    这个 workbook 对象代表这个 Excel 文件，有点类似 File 对象代表一个打开的文本文件。
    要记住，example.xlsx 需要在当前工作目录，你才能处理它。
    可以导入 os，使用函数 os.getcwd()弄清楚当前工作目录是什么，
    并使用 os.chdir()改变当前工作目录。

"""




# wb = openpyxl.load_workbook('3018024266901.xlsx')

# censuspopdata.xlsx

# print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'>

# DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
# print(wb.get_sheet_names())

print(wb.sheetnames) #['Sheet1', 'Sheet2', 'Sheet3']

# DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
#  sheet3 = wb.get_sheet_by_name('Sheet3')

# sheet3 = wb['Sheet3']

# print(sheet3) #<Worksheet "Sheet3">
# print(type(sheet3)) #<class 'openpyxl.worksheet.worksheet.Worksheet'>

# print(sheet3.title)
# sheet = wb.get_active_sheet()
# sheet = wb.active  #<Worksheet "Sheet1">


sheet = wb['Sheet1']
# print(sheet)
# print(sheet['A1'].value)
#
# print(sheet.cell(row = 3,column = 2))
# print(sheet.cell(row = 4,column = 1).value)
# print(sheet.max_row)
# print(sheet.min_row)
# print(sheet.max_col)


#列字母和数字之间的转换

# print(get_column_letter(1))
# print(get_column_letter(sheet.max_row))
# print(column_index_from_string('A'))





# print(sheet['A3':'C3'])   sheet['A3':'D3']
# for rowOfCellObjs in sheet['A3':'D3']:
#     for cellObj in  rowOfCellObjs:
#         print(cellObj.coordinate,cellObj.value)
#
#     print('END OF ROW')
#
# print(sheet.columns[0])



#读取表格的数据


# print(sheet)
# countyData = {}

for row in range(3,sheet.max_row +1):

    job = sheet['B' + str(row)].value
    name = sheet['C' + str(row)].value
    phone  =  sheet['D' + str(row)].value
    print(job)
    print(name)
    print(phone)
    # print(sheet['E' + str(row)].value)
    #
    # print(sheet['F' + str(row)].value)
    #
    # print(sheet['G' + str(row)].value)
    # print(sheet['H' + str(row)].value)
    #
    # print(sheet['I' + str(row)].value)
    # print(sheet['J' + str(row)].value)
    # print(sheet['K' + str(row)].value)
    #
    # print(sheet['L' + str(row)].value)
    #
    # print(sheet['M' + str(row)].value)
    # print(sheet['N' + str(row)].value)
    # print(sheet['O' + str(row)].value)








