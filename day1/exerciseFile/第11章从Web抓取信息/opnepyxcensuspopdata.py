# !/usr/bin/python3
# -*- coding: UTF-8 -*-
import pprint

__author__ = 'joedd'


import  openpyxl,os

from openpyxl.utils.cell import get_column_letter,column_index_from_string
# openpyxl 模块让 Python 程序能读取和修改 Excel 电子表格文件。



wb = openpyxl.load_workbook('censuspopdata.xlsx')



# print(wb.sheetnames) #['Sheet1', 'Sheet2', 'Sheet3']

countyData = {'AK': {'Aleutians East':{'pop': 3141, 'tracts': 1},
                    'Aleutians West': {'pop': 5561, 'tracts': 2},
                         'Anchorage': {'pop': 291826, 'tracts': 55},
                            'Bethel': {'pop': 17013, 'tracts': 3},
                       'Bristol Bay': {'pop': 997, 'tracts': 1}
                     }
              }


sheet = wb['Population by Census Tract']



for row in range(2,sheet.max_row +1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop  =  sheet['D' + str(row)].value
    # print(state)
    # print(county)
    # print(pop)
    #确保key存在
    countyData.setdefault(state,{})
    countyData[state].setdefault(county,{'pop': 0, 'tracts': 0})
    #加1
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)





#写入文件
print("写入的结果")
print('Writing results...')
resultFile = open('census2018.py', 'w')
#pprint.pformat()函数产生一个字符串，它本身就是格式化好的、有效的 Python代码。
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')

# print(wb.sheetnames) #['Sheet1', 'Sheet2', 'Sheet3']




#读取
path = os.chdir('/Users/joedd/Documents/PycharmProjects/day1/exerciseFile/第11章从Web抓取信息/census2018.py')
print(path)
















